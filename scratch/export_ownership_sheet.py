import copy
import os
import sys
from datetime import date, datetime

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app import app, db, OwnershipTrip, OWNERSHIP_SHEET_PATH


OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs", "ownership")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "Ownership Sheet Populated.xlsx")

STATUS_TO_SHEET = {
    "complete": "COMPLETE",
    "ongoing": "ONGOING",
    "pending": "PENDING",
    "notrequired": "Not Required",
}


def sheet_status(value):
    return STATUS_TO_SHEET.get((value or "").strip().lower(), value or "")


def write_cell(ws, row, col, value, template_row=3):
    cell = ws.cell(row, col)
    template = ws.cell(template_row, col)
    if template.has_style:
        cell._style = copy.copy(template._style)
    if template.number_format:
        cell.number_format = template.number_format
    cell.font = copy.copy(template.font)
    cell.fill = copy.copy(template.fill)
    cell.border = copy.copy(template.border)
    cell.alignment = copy.copy(template.alignment)
    cell.value = value


def export_sheet():
    with app.app_context():
        trips = OwnershipTrip.query.order_by(
            OwnershipTrip.start_date.desc().nullslast(),
            OwnershipTrip.source_row.asc().nullslast(),
            OwnershipTrip.created_at.desc(),
        ).all()

        wb = openpyxl.load_workbook(OWNERSHIP_SHEET_PATH)
        ws = wb.active

        for row in range(3, ws.max_row + 1):
            for col in range(1, 20):
                ws.cell(row, col).value = None

        for offset, trip in enumerate(trips, start=3):
            values = [
                None,
                trip.guest_name,
                trip.pax,
                trip.destination,
                trip.start_date,
                sheet_status(trip.proposal_status),
                sheet_status(trip.flights_status),
                sheet_status(trip.visa_status),
                sheet_status(trip.hotels_status),
                sheet_status(trip.sector_tickets_status),
                sheet_status(trip.sightseeing_status),
                sheet_status(trip.insurance_status),
                sheet_status(trip.traveling_status),
                sheet_status(trip.travefy_task_list_status),
                sheet_status(trip.trip_feedback_form_status),
                (trip.owner or "").upper() if trip.owner else "",
                trip.last_status_update_date,
                trip.latest_update,
                trip.present_work_assigned_to,
            ]
            for col, value in enumerate(values, start=1):
                write_cell(ws, offset, col, value)

        if trips:
            ws.max_row
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        wb.save(OUTPUT_PATH)
        db.session.remove()
        print(OUTPUT_PATH)
        print(len(trips))


if __name__ == "__main__":
    export_sheet()
